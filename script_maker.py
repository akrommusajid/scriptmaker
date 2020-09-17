from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from jinja2 import Template
from pprint import pprint

class scriptDB:
    def __init__(self, filedb):
        self.wb = load_workbook(filedb)
    
    def integrationDB(self):
        ws = self.wb["integration"]
        data_header = [ cell.value for cell in ws[1] ]
        max_row = ws.max_row
        integration_data = list()
        for row in range(2, max_row+1):
            data = dict()
            data_field = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                data[data_header[x]] = data_field[x]
            integration_data.append(data)
        return integration_data
    
    def ospfDB(self):
        ws = self.wb["ospf"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        ospf_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_field = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                if data_header[x] == "network" or data_header[x] == "wildcard" or data_header[x] == "description":
                    data[data_header[x]] = data_field[x].splitlines()
                else:
                    data[data_header[x]] = data_field[x]
            ospf_data.append(data)
        return ospf_data

    def bgpDB(self):
        ws = self.wb["bgp"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        bgp_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_field = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                data[data_header[x]] = data_field[x]
            bgp_data.append(data)
        return bgp_data
    
    def bgpConsolidationDB(self):
        ws = self.wb["bgp_as_consolidation"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        bgp_consol_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_field = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                #print(data_field[x])
                if "\n" in str(data_field[x]):
                    data[data_header[x]] = data_field[x].split("\n")
                    if data_header[x] == "vrf_redistribute":
                        data[data_header[x]] = [ tuple(x.split(",")) for x in data[data_header[x]]]
                    elif data_header[x] == "neighbor_as":
                        data[data_header[x]] = [ int(x) for x in data[data_header[x]] ]
                else:
                    data[data_header[x]] = data_field[x]
            bgp_consol_data.append(data)
        return bgp_consol_data
    
    def vrfDB(self):
        ws = self.wb["vrf"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        vrf_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_record = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                if data_header[x] == "rt_export" or data_header[x] == "rt_import":
                    data[data_header[x]] = data_record[x].splitlines()
                else:
                    data[data_header[x]] = data_record[x]
            vrf_data.append(data)
        return vrf_data

    def portMigrationDB(self):
        ws = self.wb["port_migration"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        port_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_record = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                if data_header[x] == "member_port_A" or data_header[x] == "vlan_port_A":
                    if data_record[x] != None:
                        data[data_header[x]] = data_record[x].splitlines()
                    else:
                        data[data_header[x]] = None
                else:
                    data[data_header[x]] = data_record[x]
            port_data.append(data)
        return port_data
    
    def portChannelDB(self):
        ws = self.wb["portchannel"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        port_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_record = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                if data_header[x] == "node_A_interface" or data_header[x] == "node_B_interface":
                    data[data_header[x]] = data_record[x].splitlines()
                else:
                    data[data_header[x]] = data_record[x]
            port_data.append(data)
        return port_data
    
    def staticDB(self):
        ws = self.wb["static"]
        data_header = [ cell.value for cell in ws[1] ]
        current_row = ws.max_row
        static_data = list()
        for row in range(2, current_row+1):
            data = dict()
            data_record = [ cell.value for cell in ws[row] ]
            for x in range(len(data_header)):
                data[data_header[x]] = data_record[x]
            static_data.append(data)
        return static_data

class scriptGenerator:
    def __init__(self):
        self.interconnect_service_instance_template = '''
        interface {{ data_field["node_A_interface"] }} 
         {%- if data_field["mtu"] != '' %}
         mtu {{ data_field["mtu"] }}
         {%- endif %}
         description {{ data_field["node_A"] }}_{{ data_field["node_A_interface"] }}_to_{{ data_field["node_B"] }}_{{ data_field["node_B_interface"] }}
         service-policy input map_backbone_ingress
         no shutdown
         service instance {{ data_field["service_instance"] }} ethernet
          encapsulation dot1q {{ data_field["dot1q"] }}
          rewrite ingress tag pop 1 symmetric
          bridge-domain {{ data_field["bridge_domain"] }}
        !
        {%- if data_field["interface_vlan"] != None %}
        interface {{ data_field["interface_vlan"] }}
         {%- if data_field["vrf"] != 'Global' %}
         description {{ data_field["node_A"] }}_to_{{ data_field["node_B"] }}_vrf_{{ data_field["vrf"] }}
         vrf forwarding {{ data_field["vrf"] }}
         {%- else %}
         description {{ data_field["node_A"] }}_to_{{ data_field["node_B"] }}
         {%- endif %}
         ip address {{ data_field["node_A_ip"] }} {{ data_field["netmask"] }}
         no shutdown
        {%- endif %}
        '''
        # self.interconnect_interface_vlan_template = '''

        # '''
        self.ospf_template = '''
        interface {{ data_field["node_A_interface"] }}
         {%- if data_field["mtu"] != '' %}
         ip mtu {{ data_field["mtu"] }}
         {%- endif %}
         ip ospf {{ data_field["process_id"] }} area {{ data_field["area"] }}
         {%- if data_field["network_type"] != '' %}
         ip ospf network {{ data_field["network_type"] }}
         {%- endif %}
        '''
        self.bgp_main_template = '''
        router bgp {{ data_field["node_A_as"] }}
        '''
        self.bgp_template = '''
        {%- if data_field["vrf"] != "global" %}
         address-family ipv4 vrf {{ data_field["vrf"] }}
          neighbor {{ data_field["node_B_ip"] }} remote-as {{ data_field["node_B_as"] }}
          neighbor {{ data_field["node_B_ip"] }} activate
          neighbor {{ data_field["node_B_ip"] }} description to_{{ data_field["node_B"] }}_{{ data_field["vrf"] }}
          neighbor {{ data_field["node_B_ip"] }} send-community extended
         exit-address-family
        {%- else %}
         neighbor {{ data_field["node_B_ip"] }} remote-as {{ data_field["node_B_as"] }}
         {%- if data_field["node_A_local_as"] == "yes" %} 
         neighbor {{ data_field["node_B_ip"] }} local-as {{ data_field["node_B_as"] }}
         neighbor {{ data_field["node_B_ip"] }} update-source loopback0
         {%- endif %}
         neighbor {{ data_field["node_B_ip"] }} description to_{{ data_field["node_B"] }}
         neighbor {{ data_field["node_B_ip"] }} fall-over bfd
        {%- endif %}
        {%- if data_field["vpnv4"] == "yes" %}
         address-family vpnv4
          neighbor {{ data_field["node_B_ip"] }} activate
          neighbor {{ data_field["node_B_ip"] }} send-community both
         exit-address-family
        {%- endif %}
        '''
        self.bgp_consol_template = '''
        no router bgp {{ data_field["node_A_old_as"] }}
        router bgp {{ data_field["node_A_as"] }}
         {%- if data_field["node_A_routerid"] != None %}
         bgp router-id {{ data_field["node_A_routerid"] }}
         no bgp default ipv4-unicast
         {%- endif %}
         {%- for neighbor, neighbor_ip, remote_as, local_as in neighbors %}
         neighbor {{ neighbor_ip }} remote-as {{ remote_as }}
         neighbor {{ neighbor_ip }} description {{ neighbor }}
         neighbor {{ neighbor_ip }} fall-over bfd
         {%- if data_field["node_A_as"] == remote_as %}
         neighbor {{ neighbor_ip }} update-source loopback0
         {%- elif local_as == "yes" %}
         neighbor {{ neighbor_ip }} local-as {{ remote_as }}
         neighbor {{ neighbor_ip }} update-source loopback0
         {%- endif %}
         {%- endfor %}
         address-family vpnv4
         {%- for neighbor_ip, vpnv4 in vpnv4_address_family %}
         {%- if vpnv4 == "yes" %}
          neighbor {{ neighbor_ip }} activate
          neighbor {{ neighbor_ip }} send-community extended
         {%- endif %}
         {%- endfor %}
         exit-address-family
         !
         {%- for vrf, redistribute in ipv4_address_family %}
         address-family ipv4 vrf {{ vrf }}
         {%- for protocol in redistribute %}
          redistribute {{ protocol }}
         {%- endfor %}
         exit-address-family
         !
         {%- endfor %}
        '''
        self.bgp_consol_huawei_template = '''
        undo bgp {{ data["node_A_old_as"] }}
        bgp {{ data["node_A_as"] }}
         {%- if data["node_A_routerid"] != None %}
         router-id {{ data["node_A_routerid"] }}
         {%- endif %}
         {%- for neighbor, ip, remote_as, local_as in neighbors %}
         peer {{ ip }} as-number {{ remote_as }}
         pper {{ ip }} description {{ neighbor }}
         peer {{ ip }} bfd enable
         peer {{ ip }} bfd min-tx-interval 100 min-rx-interval 100 detect-multiplier 3
         {%- if data["node_A_as"] == remote_as %}
         peer {{ ip }} connect-interface Loopback0
         {%- endif %}
         {%- endfor %}
        
         ipv4-family vpnv4
          policy vpn-target
          {%- for ip, vpnv4 in vpnv4_address_family %}
          {%- if vpnv4 == "yes" %}
          peer {{ ip }} enable
          {%- endif %}
          {%- endfor %}
         
         {%- for vrf, redistribute in ipv4_address_family %}
         ipv4-family vpn-instance {{ vrf }}
          {%- for protocol in redistribute %}
          {%- if protocol == "connected" %}
          import-route direct
          {%- else %}
          import-route {{ protocol }}
          {%- endif %}
          {%- endfor %}
        {%- endfor %}
        '''
        self.vrf_template = '''
        vrf definition {{ data["vrf_name"]}}
        {%- if data["rd"] != None %}
         rd {{ data["rd"] }}
        {%- endif %}
         address-family ipv4
         {%- for export in data["rt_export"] %}
          route-target export {{ export }}
         {%- endfor %}
         {%- for import in data["rt_import"] %}
          route-target import {{ import }}
         {%- endfor %}
         exit-address-family
        '''
        self.vrf_huawei_template = '''
        ip vpn-instance {{ data["vrf_name"] }}
         ipv4-family
          {%- for export in data["rt_export"] %}
          vpn-target {{ export }} export-extcommunity
          {%- endfor %}
          {%- for import in data["rt_import"] %}
          vpn-target {{ import }} import-extcommunity
          {%- endfor %}
        '''
        self.port_template = '''
        interface {{ data["port_A"] }}
         shutdown
        {%- if data["member_port_A"] != None %}
        {%- for member in data["member_port_A"] %}
        interface {{ member }}
         shutdown
        {%- endfor %}
        {% endif %}
        {%- if data["vlan_port_A"] != None %}
        {%- for member in data["vlan_port_A"] %}
        interface {{ member }}
         shutdown
        {%- endfor %}
        {%- endif %}
        '''
        self.portchannel_template_huawei = '''
        interface eth-trunk{{ data["node_A_portchannel"] }}
         description {{ data["node_A"] }}_to_{{ data["node_B"] }}
         mode lacp-static
         load-balance packet-all
         statistic enable

        {%- for portA, portB in portchannel %}
        interface {{ portA }}
         description {{ data["node_A"] }}_{{ portA }}_to_{{ data["node_B"] }}_{{ portB }}
         undo shutdown
         eth-trunk {{ data["node_A_portchannel"] }}
        {%- endfor %}
        '''
        self.interconnect_subif_template_huawei = '''
        interface {{ data["node_A_interface"] }}.{{ data["sub_if"] }}
         vlan-type dot1q {{ data["dot1q"] }}
         mtu {{ data["mtu"] }}
         {%- if data["remark"] != None %}
         description {{ data["node_A"] }}_to_{{ data["node_B"] }}_{{ data["vrf"] }} {{ data["remark"] }}
         {%- else %}
         description {{ data["node_A"] }}_to_{{ data["node_B"] }}_{{ data["vrf"] }}
         {%- endif %}
         {%- if data["vrf"] != "global" %}
         ip binding vpn-instance {{ data["vrf"] }}
         {%- endif %}
         ip address {{ data["node_A_ip"] }} {{ data["netmask"] }}
         statistic enable
         trust upstream default
        '''
        self.bgp_as_template_huawei = '''
        bgp {{ data["node_A_as"] }}
        '''
        self.bgp_template_huawei = '''
        {%- if data["vrf"] != "global" %}
         ipv4-family vpn-instance {{ data["vrf"] }}
          peer {{ data["node_B_ip"] }} as-number {{ data["node_B_as"] }}
          peer {{ data["node_B_ip"] }} description {{ data["node_B"] }}_{{ data["vrf"] }}
          peer {{ data["node_B_ip"] }} bfd enable
          peer {{ data["node_B_ip"] }} bfd min-tx-interval 100 min-rx-interval 100 detect-multiplier 3
          {%- if data["auth"] != None %}
          peer {{ data["node_B_ip"] }} password cipher {{ data["auth"] }}
          {%- endif %}
        {%- else %}
          peer {{ data["node_B_ip"] }} as-number {{ data["node_B_as"] }}
          peer {{ data["node_B_ip"] }} description {{ data["node_B"] }}
          peer {{ data["node_B_ip"] }} bfd enable
          peer {{ data["node_B_ip"] }} bfd min-tx-interval 100 min-rx-interval 100 detect-multiplier 3
        {%- endif %}
        {%- if data["node_A_as"] == data["node_B_as"] %}
          peer {{ data["node_B_ip"] }} connect-interface Loopback0
        {%- endif %}
        {%- if data["node_A_local_as"] == "yes" %}
          peer {{ data["node_B_ip"] }} connect-interface Loopback0
          peer {{ data["node_B_ip"] }} local-as {{ data["node_B_as"] }}
        {%- endif %}
        {%- if data["vpnv4"] == "yes" %}
         ipv4-family vpnv4
          peer {{ data["node_B_ip"] }} enable
        {%- endif %}
          quit
        '''
        self.staticroute_template_huawei = '''
        {%- if data["vrf"] != "Global" %}
        ip route-static vpn-instance {{ data["vrf"] }} {{ data["network"] }} {{ data["netmask"] }} {{ data["nexthop"] }} description {{ data["description"] }}
        {%- endif %}
        '''
        self.ospf_template_huawei = '''
        {%- if data["network_type"] == "point-to-point" %}
        interface {{ data["node_A_interface"] }}
         ospf network-type p2p
        {%- endif %}
        {%- if data["vrf"] != "Global" %}
        ospf {{ data["process_id"] }} router-id {{ data["node_A_routerid"] }} vpn-instance {{ data["vrf"] }}
        {%- endif %} 
         area {{ data["area"] }}
        {%- for network, wildcard, description in networks %}
          network {{ network }} {{ wildcard }} description {{ description }}
        {%- endfor %}
        '''
    
    def staticroute(self, staticroute):
        if staticroute["node_A_vendor"] == "huawei":
            staticroute_template = Template(self.staticroute_template_huawei)
            result = staticroute_template.render(data=staticroute)
            return result

    def interconnect(self, interconnect):
        if interconnect["node_A_vendor"] == "huawei":
            interconnect_template = Template(self.interconnect_subif_template_huawei)
            result = interconnect_template.render(data=interconnect)
            return result
        else:
            interconnect_template = Template(self.interconnect_service_instance_template)
            result = interconnect_template.render(data_field=interconnect)
            return result
    
    def interface_vlan(self, interface_vlan):
        interface_vlan_template = Template(self.interconnect_interface_vlan_template)
        result = interface_vlan_template.render(data_field=interface_vlan)
        return result
    
    def ospf(self, ospf):
        if ospf["node_A_vendor"] == "huawei":
            ospf_template = Template(self.ospf_template_huawei)
            data = {
                "data" : ospf,
                "networks" : tuple(zip(ospf["network"], ospf["wildcard"], ospf["description"]))
            }
            result = ospf_template.render(**data)
            return result
        else:
            ospf_template = Template(self.ospf_template)
            result = ospf_template.render(data_field=ospf)
            return result
    
    def bgp_main(self, bgp):
        if bgp["node_A_vendor"] == "huawei":
            bgp_as_template = Template(self.bgp_as_template_huawei)
            result = bgp_as_template.render(data=bgp)
            return result
        else:
            bgp_main_template = Template(self.bgp_main_template)
            result = bgp_main_template.render(data_field=bgp)
            return result
    
    def bgp(self, bgp):
        if bgp["node_A_vendor"] == "huawei":
            bgp_template = Template(self.bgp_template_huawei)
            result = bgp_template.render(data=bgp)
            return result
        else:
            bgp_template = Template(self.bgp_template)
            result = bgp_template.render(data_field=bgp)
            return result
    
    def bgp_consol(self, bgp_consol):
        if bgp_consol["node_A_vendor"] == "huawei":
            bgp_consolidation_template = Template(self.bgp_consol_huawei_template)
            data = {
                "data" : bgp_consol,
                "vpnv4_address_family" : tuple(zip(bgp_consol["neighbor_ip"], bgp_consol["vpnv4"])),
                "neighbors" : tuple(zip(bgp_consol["node_A_neighbor"], bgp_consol["neighbor_ip"], bgp_consol["neighbor_as"], bgp_consol["local_as"]))
            }
            if  bgp_consol["vrf"] != None and bgp_consol["vrf_redistribute"] != None:
                data["ipv4_address_family"] = tuple(zip(bgp_consol["vrf"], bgp_consol["vrf_redistribute"]))
            result = bgp_consolidation_template.render(**data)
            return result
        else:
            bgp_consolidation_template = Template(self.bgp_consol_template)
            #pprint(bgp_consol)
            data = {
                "data_field" : bgp_consol,
                "vpnv4_address_family" : tuple(zip(bgp_consol["neighbor_ip"], bgp_consol["vpnv4"])),
                "neighbors" : tuple(zip(bgp_consol["node_A_neighbor"], bgp_consol["neighbor_ip"], bgp_consol["neighbor_as"], bgp_consol["local_as"]))
            }
            if  bgp_consol["vrf"] != None and bgp_consol["vrf_redistribute"] != None:
                data["ipv4_address_family"] = tuple(zip(bgp_consol["vrf"], bgp_consol["vrf_redistribute"]))
            # vpnv4_address_family = tuple(zip(bgp_consol["neighbor_ip"], bgp_consol["vpnv4"]))
            # ipv4_address_family = tuple(zip(bgp_consol["vrf"], bgp_consol["redistribute"]))
            # neighbors = tuple(zip(bgp_consol["node_A_neighbor"], bgp_consol["neighbor_ip"], bgp_consol["neighbor_as"]))
            result = bgp_consolidation_template.render(**data)
            return result
    
    def vrf(self, vrf):
        if vrf["node_vendor"] == "huawei":
            vrf_template = Template(self.vrf_huawei_template)
            result = vrf_template.render(data=vrf)
            return result
        else:
            vrf_template = Template(self.vrf_template)
            result = vrf_template.render(data=vrf)
            return result
    
    def port_migration(self, port_migration):
        port_template = Template(self.port_template)
        result = port_template.render(data=port_migration)
        return result
    
    def portchannel(self, portchannel):
        if portchannel["node_A_vendor"] == "huawei":
            portchannel_template = Template(self.portchannel_template_huawei)
            data = {
                "data" : portchannel,
                "portchannel" : tuple(zip(portchannel["node_A_interface"], portchannel["node_B_interface"]))
            }
            result = portchannel_template.render(**data)
            return result

class xportMop:
    def __init__(self, interconnect=[], ospf=[], bgp=[], bgp_consolidation=[], vrf=[], port_migration=[], portchannel=[], staticroute=[]):
        self.wb = Workbook()
        self.interconnect = interconnect
        self.ospf = ospf
        self.bgp = bgp
        self.bgp_consolidation = bgp_consolidation
        self.vrf = vrf
        self.port_migration = port_migration
        self.portchannel = portchannel
        self.staticroute = staticroute
        #self.interface_vlan = interface_vlan
    
    def steps(self):
        ws = self.wb.create_sheet("Steps")
        ws["A1"] = "Phase"
        ws["B1"] = "Activity"
        ws["C1"] = "Risk"
        index_phase = ["A", "B", "C", "D", "E", "F"]
        idx = 0
        current_phase = index_phase[idx]
        current_row = ws.max_row+1
        result = list()
        len_column_B = 0
        len_column_C = 0
        self.node = list()
        if len(self.portchannel) > 0:
            node_A = [ x["node_A"] for x in self.portchannel ]
            self.node += node_A
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Add Port-channel"
            current_row += 1
            for index, data in enumerate(self.portchannel):
                record = dict()
                ws["A%s" % current_row] = record["phase"] = "%s%s" % (current_phase, index+1)
                for interface in data["node_A_interface"]:
                    ws["B%s" % current_row] = record["activity"] = "Add interface %s to link aggregation id %s on %s" % (interface, data["node_A_portchannel"], data["node_A"])
                    #print(len(ws["B%s" % current_row].value))
                    if len_column_B < len(record["activity"]):
                        ws.column_dimensions["B"].width = len_column_B = len(record["activity"])
                    ws["C%s" % current_row] = "No Downtime"
                    if len_column_C < len(ws["C%s" % current_row].value):
                        ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                    current_row += 1
                result.append(record)
            idx += 1
            current_phase = index_phase[idx]        
        if len(self.interconnect) > 0:
            node_A = [ x["node_A"] for x in self.interconnect ]
            #node_B = [ x["node_B"] for x in self.interconnect ]
            self.node += node_A
            #self.node += node_B
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Integration"
            for index, data in enumerate(self.interconnect):
                field = dict()
                current_row += 1
                ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = field["activity"] = "Create P2P %s to %s vrf %s" % (data["node_A"], data["node_B"], data["vrf"])
                #print(len(ws["B%s" % current_row].value))
                if len_column_B < len(field["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(field["activity"])
                ws["C%s" % current_row] = "No Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(field)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1
        if len(self.staticroute) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Static route"
            node_A = [ x["node_A"] for x in self.staticroute ]
            #node_B = [ x["node_B"] for x in self.ospf ]
            self.node += node_A
            node_A = set(node_A)
            #self.node += node_B
            for index, node in enumerate(node_A):
                field = dict()
                current_row += 1
                ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = field["activity"] = "Static route on %s" % node
                field["node_A"] = node
                if len_column_B < len(field["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(field["activity"])
                ws["C%s" % current_row] = "No Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(field)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1            
        if len(self.ospf) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Enable OSPF"
            node_A = [ x["node_A"] for x in self.ospf ]
            #node_B = [ x["node_B"] for x in self.ospf ]
            self.node += node_A
            #self.node += node_B
            for index, data in enumerate(self.ospf):
                field = dict()
                current_row += 1
                ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = field["activity"] = "Enable OSPF %s to %s" % (data["node_A"], data["node_B"])
                if len_column_B < len(field["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(field["activity"])
                ws["C%s" % current_row] = "No Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(field)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1
        if len(self.bgp) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Create BGP"
            node_A = [ x["node_A"] for x in self.bgp ]
            #node_B = [ x["node_B"] for x in self.bgp ]
            self.node += node_A
            #self.node += node_B
            for index, data in enumerate(self.bgp):
                field = dict()
                current_row += 1
                if data["node_A_as"] != data["node_B_as"]:
                    if data["vrf"] != "Global":
                        ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                        ws["B%s" % current_row] = field["activity"] = "Enable eBGP %s to %s vrf %s" % (data["node_A"], data["node_B"], data["vrf"])
                        if len_column_B < len(field["activity"]):
                            ws.column_dimensions["B"].width = len_column_B = len(field["activity"])
                    else:
                        ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                        ws["B%s" % current_row] = field["activity"] = "Enable BGP %s to %s vrf %s" % (data["node_A"], data["node_B"], data["vrf"])
                        if len_column_B < len(field["activity"]):
                            ws.column_dimensions["B"].width = len_column_B = len(field["activity"])      
                ws["C%s" % current_row] = "No Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)  
                result.append(field)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1  
        if len(self.vrf) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Add export and import VRF"
            node_A = [ x["node"] for x in self.vrf ]
            self.node += node_A
            for index, data in enumerate(self.vrf):
                record = dict()
                current_row += 1
                ws["A%s" % current_row] = record["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = record["activity"] = "Add export import vrf %s on %s" % (data["vrf_name"], data["node"])
                if len_column_B < len(record["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(record["activity"])
                ws["C%s" % current_row] = " No downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(record)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1   
        if len(self.port_migration) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "Port Migration"
            node_A = [ x["node_A"] for x in self.port_migration ]
            self.node += node_A
            for index, data in enumerate(self.port_migration):
                record = dict()
                current_row += 1
                ws["A%s" % current_row] = record["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = record["activity"] = "Port migration %s %s on %s" % (data["description"], data["port_A"], data["node_A"])
                if len_column_B < len(record["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(record["activity"])
                ws["C%s" % current_row] = " Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(record)
            idx += 1
            current_phase = index_phase[idx]
            current_row += 1
        if len(self.bgp_consolidation) > 0:
            ws["A%s" % current_row] = current_phase
            ws["B%s" % current_row] = "BGP consolidation"
            node_A = [ x["node_A"] for x in self.bgp_consolidation ]
            self.node += node_A
            for index, data in enumerate(self.bgp_consolidation):
                field = dict()
                current_row += 1
                ws["A%s" % current_row] = field["phase"] = "%s%s" % (current_phase, index+1)
                ws["B%s" % current_row] = field["activity"] = "BGP consolidation on %s" % data["node_A"]
                if len_column_B < len(field["activity"]):
                    ws.column_dimensions["B"].width = len_column_B = len(field["activity"])
                ws["C%s" % current_row] = "Downtime"
                if len_column_C < len(ws["C%s" % current_row].value):
                    ws.column_dimensions["C"].width = len_column_C = len(ws["C%s" % current_row].value)
                result.append(field)
        return result
    
    def script(self):
        script_gen = scriptGenerator()
        len_column = 0
        ws = self.wb.create_sheet("Script")
        ws["A1"] = "Phase"
        ws["B1"] = "Activity"
        column = ["C","D","E","F","G","H"]
        steps = self.steps()
        #return steps
        nodes = set(self.node)
        node_column = list()
        for idx, node in enumerate(nodes):
            column_n = dict()
            column_n["column"] = column[idx]
            ws["%s1"% column[idx]] = column_n["node"] = node
            node_column.append(column_n)
        #print(node_column)
        current_row = ws.max_row+1
        script = {
            "integration" : [],
            "ospf" : [],
            "bgp" : [],
            "bgp_consolidation" : [],
            "vrf" : [],
            "port_migration" : [],
            "portchannel" : [],
            "staticroute" : []
        }
        for step in steps:
            if "P2P" in step["activity"]:
                script["integration"].append(step)
            elif "OSPF" in step["activity"]:
                script["ospf"].append(step)
            elif "consolidation" in step["activity"]:
                script["bgp_consolidation"].append(step)
            elif "BGP" in step["activity"]:
                script["bgp"].append(step)
            elif "export import" in step["activity"]:
                script["vrf"].append(step)
            elif "Port migration" in step["activity"]:
                script["port_migration"].append(step)
            elif "link aggregation" in step["activity"]:
                script["portchannel"].append(step)
            elif "Static route" in step["activity"]:
                script["staticroute"].append(step)
        if len(self.portchannel) > 0:
            for index, data in enumerate(script["portchannel"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                portchannel_script = script_gen.portchannel(self.portchannel[index])
                #print(portchannel_script)
                portchannel_script = portchannel_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.portchannel[index]["node_A"]]
                for s in portchannel_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1
        if len(self.interconnect) > 0:
            for index, data in enumerate(script["integration"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                interconnect_script = script_gen.interconnect(self.interconnect[index])
                #print(interconnect_script)
                interconnect_script = interconnect_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.interconnect[index]["node_A"]]
                for s in interconnect_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1
        if len(self.staticroute) > 0:
            for index, data in enumerate(script["staticroute"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                node_name = data["node_A"]
                #print(node_name)
                staticroute_script = str()
                for static in self.staticroute:
                    #print(static)
                    node_A = static["node_A"]
                    if node_A == node_name:
                        #print(node_name)
                        #print(node_A)
                        staticroute_script += script_gen.staticroute(static)
                    else:
                        continue
                print(staticroute_script)
                staticroute_script = staticroute_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == data["node_A"]]
                for s in staticroute_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1                
        if len(self.ospf) > 0:
            for index, data in enumerate(script["ospf"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                ospf_script = script_gen.ospf(self.ospf[index])
                print(ospf_script)
                ospf_script = ospf_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.ospf[index]["node_A"]]
                for s in ospf_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    current_row += 1
        if len(self.bgp) > 0:
            node_name = None
            for index, data in enumerate(script["bgp"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                node_A = self.bgp[index]["node_A"]
                bgp_script = str()
                if node_A != node_name:
                    node_name = node_A
                    bgp_script = script_gen.bgp_main(self.bgp[index])
                    #print(script)
                    #bgp_script += script
                bgp_script += script_gen.bgp(self.bgp[index])
                #print(bgp_script)
                bgp_script = bgp_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.bgp[index]["node_A"]]
                for s in bgp_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1
        if len(self.vrf) > 0:
            for index, data in enumerate(script["vrf"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                else:
                    ws.column_dimensions["B"].width = len_column
                vrf_script = script_gen.vrf(self.vrf[index])
                vrf_script = vrf_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.vrf[index]["node"]]
                for s in vrf_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    #print(len_column)
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1
        if len(self.port_migration) > 0:
            for index, data in enumerate(script["port_migration"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                else:
                    ws.column_dimensions["B"].width = len_column
                vrf_script = script_gen.port_migration(self.port_migration[index])
                vrf_script = vrf_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.port_migration[index]["node_A"]]
                for s in vrf_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    #print(len_column)
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1
        if len(self.bgp_consolidation) > 0:
            #pprint(self.bgp_consolidation)
            for index, data in enumerate(script["bgp_consolidation"]):
                ws["A%s" % current_row] = data["phase"]
                ws["B%s" % current_row] = data["activity"]
                if len_column < len(data["activity"]):
                    ws.column_dimensions["B"].width = len_column = len(data["activity"])
                bgp_consol_script = script_gen.bgp_consol(self.bgp_consolidation[index])
                bgp_consol_script = bgp_consol_script.splitlines()
                node_col = [ x["column"] for x in node_column if x["node"] == self.bgp_consolidation[index]["node_A"]]
                #print(node_col[0])
                for s in bgp_consol_script:
                    ws["%s%s" % (node_col[0], current_row)] = s
                    #print(len_column)
                    if len_column < len(s):
                        ws.column_dimensions["%s" % node_col[0]].width = len_column = len(s)
                    else:
                        ws.column_dimensions["%s" % node_col[0]].width = len_column
                    current_row += 1  
        return script

    def save(self, name):
        self.wb.save(name)

def main():
    filedb = "mop_db.xlsx"
    database = scriptDB(filedb)
    interconnect_data = database.integrationDB()
    ospf_data = database.ospfDB()
    bgp_data = database.bgpDB()
    bgp_consol_data = database.bgpConsolidationDB()
    port_data = database.portMigrationDB()
    vrf_data = database.vrfDB()
    portchannel_data = database.portChannelDB()
    staticroute_data = database.staticDB()
    print("=========================================== VRF =================================================================")
    pprint(vrf_data)
    print("=========================================== Portchannel =========================================================")
    pprint(portchannel_data)
    print("=========================================== Integration =========================================================")
    pprint(interconnect_data)
    print("=========================================== Staticroute =========================================================")
    pprint(staticroute_data)
    print("=========================================== OSPF ================================================================")
    pprint(ospf_data)
    print("=========================================== BGP =================================================================")
    pprint(bgp_consol_data)
    pprint(bgp_data)
    print("=========================================== Port migration ======================================================")
    pprint(port_data)
    print("=========================================== Generate script =====================================================")
    # for data in interconnect_data:
    #     script1 = scriptGenerator().interconnect(data)
    #     print(script1)
    #     script2 = scriptGenerator().interface_vlan(data)
    #     print(script2)
    # for data in ospf_data:
    #     script3 = scriptGenerator().ospf(data)
    #     print(script3)
    # for data in bgp_consol_data:
    #     script4 = scriptGenerator().bgp_consol(data)
    #     print(script4)
    create_mop = xportMop(  interconnect=interconnect_data, 
                            ospf=ospf_data, bgp=bgp_data, 
                            bgp_consolidation=bgp_consol_data, 
                            vrf=vrf_data,
                            port_migration=port_data,
                            portchannel=portchannel_data,
                            staticroute=staticroute_data)
    print("=========================================== Create MoP =========================================================")
    pprint(create_mop.script())
    #create_mop.script()
    create_mop.save("MoP.xlsx")
    
if __name__ == "__main__":
    main()